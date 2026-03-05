from flask import Flask, request, jsonify, Response, send_from_directory, send_file
import requests, time, json, os, re, io
from collections import Counter

app = Flask(__name__, static_folder="static")

ALPHABET = "abcdefghijklmnopqrstuvwxyz"
NUMBERS  = "0123456789"
ALPHANUM = ALPHABET + NUMBERS

BASE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
REDDIT_HEADERS = {"User-Agent": "KeywordResearchTool/1.0 (local research app)", "Accept": "application/json"}

GL_MAP = {
    "us":"en","gb":"en","au":"en","ca":"en","ie":"en",
    "de":"de","at":"de","ch":"de",
    "fr":"fr","be":"fr",
    "es":"es","mx":"es","ar":"es",
    "it":"it","nl":"nl","pt":"pt","br":"pt",
    "pl":"pl","ru":"ru","jp":"ja","kr":"ko","cn":"zh-CN",
    "in":"hi","ae":"ar","sa":"ar",
}

def google_headers():
    return BASE_HEADERS.copy()

# Fix N: guard against non-string items in autocomplete arrays
def get_autocomplete(query, hl="en", gl="us"):
    try:
        r = requests.get("https://suggestqueries.google.com/complete/search",
            params={"client":"firefox","q":query,"hl":hl,"gl":gl}, timeout=10)
        data = r.json()
        return [s for s in data[1] if isinstance(s, str)]
    except: return []

def get_bing_autocomplete(query):
    try:
        r = requests.get("https://api.bing.com/osjson.aspx",
            params={"query":query,"language":"en-US"}, timeout=10,
            headers={"User-Agent": BASE_HEADERS["User-Agent"]})
        data = r.json()
        return data[1] if len(data) > 1 else []
    except: return []

# Fix O: guard against bare strings or single-element lists in YouTube suggestions
def get_youtube_autocomplete(query):
    try:
        r = requests.get("https://suggestqueries.google.com/complete/search",
            params={"client":"youtube","q":query,"hl":"en","gl":"us"}, timeout=10,
            headers={"User-Agent": BASE_HEADERS["User-Agent"]})
        data = r.json()
        if len(data) < 2: return []
        return [item[0] for item in data[1]
                if isinstance(item, list) and len(item) > 0 and isinstance(item[0], str)]
    except: return []

# Fix C: bounded cache (max 500 entries), cleared between seed runs
_google_html_cache = {}

def get_google_html(query, hl="en", gl="us"):
    key = (query.lower(), hl, gl)
    if key in _google_html_cache:
        return _google_html_cache[key]
    try:
        h = google_headers()
        h["Accept-Language"] = f"{hl},{hl[:2]};q=0.9"
        r = requests.get("https://www.google.com/search",
            params={"q":query,"hl":hl,"gl":gl}, headers=h, timeout=10)
        if len(_google_html_cache) >= 500:
            del _google_html_cache[next(iter(_google_html_cache))]
        _google_html_cache[key] = r.text
        return r.text
    except: return ""

def clear_google_cache():
    _google_html_cache.clear()

def get_people_also_ask(query, hl="en", gl="us"):
    html = get_google_html(query, hl, gl)
    results=[]; seen=set()
    for pat in [r'data-q="([^"]+)"', r'"([^"]{10,150}\?)"']:
        for m in re.findall(pat, html):
            m=m.strip()
            if not m or '?' not in m: continue
            if not (10 < len(m) < 200): continue
            if m in seen: continue
            # Filter out JS code fragments — real PAA questions don't contain these
            if any(c in m for c in ['{','}','=','&&','||',';','()','.length','function','var ','const ','let ','return ','case ']): continue
            # Must contain at least one real word of 3+ letters
            words = re.findall(r'[a-zA-Z]{3,}', m)
            if len(words) < 3: continue
            seen.add(m); results.append(m)
    return results[:10]

def get_related_searches(query, hl="en", gl="us"):
    html = get_google_html(query, hl, gl)
    results=[]; seen=set()
    for m in re.findall(r'\/search\?q=([^&"]{5,80})&(?:amp;)?sa=X', html):
        m=requests.utils.unquote(m).replace('+',' ').strip()
        if m and 5<len(m)<150 and m not in seen and m.lower()!=query.lower():
            seen.add(m); results.append(m)
    return results[:15]

def get_people_also_search(query, hl="en", gl="us"):
    html = get_google_html(query, hl, gl)
    results=[]; seen=set()
    for m in re.findall(r'\/search\?q=([^&"]{5,60})&(?:amp;)?sa=X&(?:amp;)?ved', html):
        m=requests.utils.unquote(m).replace('+',' ').strip()
        if m and 4<len(m)<100 and m not in seen and m.lower()!=query.lower():
            seen.add(m); results.append(m)
    return results[:15]

# Fix D: local pack uses its own dedicated fetch (with num=10) so it never
# gets a cached base-search page that may not contain map pack HTML
def get_local_pack(query, hl="en", gl="us"):
    try:
        h = google_headers()
        h["Accept-Language"] = f"{hl},{hl[:2]};q=0.9"
        r = requests.get("https://www.google.com/search",
            params={"q":query,"hl":hl,"gl":gl,"num":10}, headers=h, timeout=10)
        html = r.text
    except:
        html = ""
    businesses=[]
    name_patterns=[
        r'<div class="[^"]*dbg0pd[^"]*"[^>]*><span[^>]*>([^<]{3,80})</span>',
        r'aria-label="([^"]{3,80})"[^>]*class="[^"]*rllt__[^"]*"',
        r'class="[^"]*OSrXXb[^"]*"[^>]*>([^<]{3,80})<',
    ]
    found_names=[]
    for pat in name_patterns:
        for m in re.findall(pat, html):
            m=m.strip()
            skip=['Google','Search','Maps','More','Menu','Directions','Website','Call']
            if m and 3<len(m)<80 and m not in found_names and not any(s.lower()==m.lower() for s in skip):
                found_names.append(m)
        if len(found_names)>=3: break
    ratings=re.findall(r'(\d\.\d)\s*\(', html)
    reviews=re.findall(r'\((\d[\d,]*)\)', html)
    phones=re.findall(r'\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]\d{4}', html)
    for i,name in enumerate(found_names[:3]):
        businesses.append({"name":name,"rating":ratings[i] if i<len(ratings) else "",
            "reviews":reviews[i].replace(',','') if i<len(reviews) else "",
            "phone":phones[i] if i<len(phones) else "","address":"","category":""})
    return businesses[:3]

def get_google_serp_urls(query, hl="en", gl="us"):
    try:
        h = google_headers()
        h["Accept-Language"] = f"{hl},{hl[:2]};q=0.9"
        r = requests.get("https://www.google.com/search",
            params={"q":query,"hl":hl,"gl":gl,"num":10}, headers=h, timeout=12)
        html = r.text
        results = []
        for m in re.findall(r'href="/url\?q=(https?://[^&"]{10,500})&', html):
            url = requests.utils.unquote(m)
            if any(skip in url for skip in ['google.','webcache.','youtube.com/watch',
                    'accounts.google','support.google','policies.google']): continue
            results.append(url)
        if len(results) < 5:
            for m in re.findall(r'data-href="(https?://[^"]{10,400})"', html):
                if m not in results and 'google.' not in m:
                    results.append(m)
        seen=set(); unique=[]
        for url in results:
            domain = re.sub(r'^https?://(www\.)?','',url).split('/')[0]
            if domain not in seen:
                seen.add(domain); unique.append(url)
        return unique[:10]
    except Exception as e:
        print(f"Google SERP error: {e}")
        return []

def get_bing_serp_urls(query):
    try:
        h = BASE_HEADERS.copy()
        h["Referer"] = "https://www.bing.com/"
        r = requests.get("https://www.bing.com/search",
            params={"q":query,"count":10}, headers=h, timeout=12)
        html = r.text
        results = []
        for m in re.findall(r'<li class="b_algo"[^>]*>.*?<a href="(https?://[^"]{10,400})"', html, re.DOTALL):
            if any(skip in m for skip in ['bing.com','microsoft.com','msn.com']): continue
            results.append(m)
        if len(results) < 5:
            for m in re.findall(r'<a href="(https?://(?!.*bing\.com)[^"]{10,400})"[^>]*class="[^"]*tilk[^"]*"', html):
                if m not in results: results.append(m)
        seen=set(); unique=[]
        for url in results:
            domain = re.sub(r'^https?://(www\.)?','',url).split('/')[0]
            if domain not in seen:
                seen.add(domain); unique.append(url)
        return unique[:10]
    except Exception as e:
        print(f"Bing SERP error: {e}")
        return []

def extract_domain(url):
    return re.sub(r'^https?://(www\.)?','',url).split('/')[0].split('?')[0]

STOP_WORDS = {
    "the","a","an","and","or","of","to","in","for","on","at","by","is","are",
    "was","were","be","been","have","has","had","with","from","that","this",
    "it","as","not","but","can","will","do","does","did","so","if","we","you",
    "he","she","they","our","your","his","her","their","its","my","all","also",
    "more","one","get","just","about","than","up","out","into","when","what",
    "how","which","who","use","used","using","may","would","could","should",
    "any","other","new","these","those","such","no","like","been","very",
    "there","here","only","then","than","now","some","i","me","us","after",
    "before","each","between","both","through","during","while","where","why",
    "am","make","made","way","see","go","take","come","know","think","need",
    "want","back","re","ve","ll","s","t","www","com","http","https","html",
    "page","click","read","find","learn","help","support","contact","home",
    "about","privacy","terms","cookies","login","sign","subscribe","email",
}

SCHEMA_FIELD_WEIGHTS = {
    "question":9,"acceptedAnswer":8,"suggestedAnswer":7,"answerCount":0,
    "step":7,"name":8,"supply":6,"tool":6,
    "headline":9,"description":7,"keywords":9,"articleSection":6,"about":6,
    "category":8,"serviceType":9,"areaServed":7,"brand":6,"model":6,"hasOfferCatalog":6,
    "item":5,"eventName":7,"reviewBody":5,
    "alternateName":7,"disambiguatingDescription":6,"text":4,"query-input":4,
}

def _walk_jsonld(obj, results, schema_type="", depth=0):
    if depth > 8: return
    if isinstance(obj, dict):
        stype = obj.get("@type", schema_type)
        if isinstance(stype, list): stype = stype[0] if stype else ""
        for key, val in obj.items():
            if key.startswith("@"): continue
            weight = SCHEMA_FIELD_WEIGHTS.get(key, 0)
            zone   = f"schema:{stype}:{key}" if stype else f"schema:{key}"
            if isinstance(val, str):
                text = val.strip()
                if len(text) < 4 or text.startswith("http"): continue
                if weight > 0 or len(text) > 15:
                    results.append((text, max(weight, 3), zone))
            elif isinstance(val, list):
                for item in val:
                    if isinstance(item, str) and len(item.strip()) > 3:
                        results.append((item.strip(), max(weight, 3), zone))
                    elif isinstance(item, dict):
                        _walk_jsonld(item, results, stype, depth+1)
            elif isinstance(val, dict):
                _walk_jsonld(val, results, stype, depth+1)
    elif isinstance(obj, list):
        for item in obj:
            _walk_jsonld(item, results, schema_type, depth+1)

# Fix Q: parse JSON-LD blocks once per page, share with both extractors
def _parse_jsonld_blocks(html):
    blocks = re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html, re.DOTALL | re.IGNORECASE
    )
    parsed = []
    for raw in blocks:
        raw = raw.strip()
        if not raw: continue
        try:
            parsed.append(json.loads(raw))
        except (json.JSONDecodeError, ValueError):
            pass
    return parsed

def extract_jsonld_texts(parsed_blocks):
    results = []
    for data in parsed_blocks:
        if isinstance(data, dict) and "@graph" in data:
            data = data["@graph"]
        _walk_jsonld(data, results)
    return results

def extract_schema_paa(parsed_blocks):
    pairs = []
    seen_q = set()

    def clean_text(s):
        if not isinstance(s, str): return ""
        s = re.sub(r'<[^>]+>', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def get_text(val):
        if isinstance(val, str): return clean_text(val)
        if isinstance(val, dict):
            for f in ("text","name","description","answerText"):
                if f in val and isinstance(val[f], str):
                    return clean_text(val[f])
        return ""

    def walk(obj, schema_type=""):
        if isinstance(obj, list):
            for item in obj: walk(item, schema_type)
            return
        if not isinstance(obj, dict): return
        stype = obj.get("@type", schema_type)
        if isinstance(stype, list): stype = stype[0] if stype else ""

        if stype in ("FAQPage","QAPage"):
            entities = obj.get("mainEntity", obj.get("mainEntityOfPage",[]))
            if isinstance(entities, dict): entities = [entities]
            for ent in (entities if isinstance(entities, list) else []):
                q = clean_text(ent.get("name", ent.get("question","")))
                raw_ans = ent.get("acceptedAnswer", ent.get("suggestedAnswer",{}))
                if isinstance(raw_ans, list): raw_ans = raw_ans[0] if raw_ans else {}
                a = get_text(raw_ans)
                if q and q not in seen_q:
                    seen_q.add(q)
                    # Fix P: raise answer limit to 500; JS controls display truncation
                    pairs.append({"question":q,"answer":a[:500],"schema_type":stype,"source_field":"mainEntity"})

        elif stype == "HowTo":
            # Fix G: emit the HowTo title itself as a keyword entry
            howto_name = clean_text(obj.get("name",""))
            if howto_name and howto_name not in seen_q:
                seen_q.add(howto_name)
                pairs.append({"question":howto_name,"answer":clean_text(obj.get("description",""))[:500],
                              "schema_type":"HowTo","source_field":"name"})
            steps = obj.get("step",[])
            if isinstance(steps, dict): steps = [steps]
            for step in (steps if isinstance(steps, list) else []):
                step_name = clean_text(step.get("name",""))
                step_text = get_text(step)
                q = step_name or step_text
                if q and q not in seen_q:
                    seen_q.add(q)
                    pairs.append({"question":q,"answer":step_text[:500] if step_text!=q else "",
                                  "schema_type":"HowTo:step","source_field":"step"})

        elif stype in ("Article","BlogPosting","NewsArticle","TechArticle","WebPage"):
            kw_field = obj.get("keywords","")
            if isinstance(kw_field, str) and kw_field:
                for kw in re.split(r'[,;|]', kw_field):
                    kw = kw.strip()
                    if kw and kw not in seen_q and len(kw)>3:
                        seen_q.add(kw)
                        pairs.append({"question":kw,"answer":"","schema_type":stype,"source_field":"keywords"})
            elif isinstance(kw_field, list):
                for kw in kw_field:
                    kw = clean_text(kw)
                    if kw and kw not in seen_q:
                        seen_q.add(kw)
                        pairs.append({"question":kw,"answer":"","schema_type":stype,"source_field":"keywords"})

        elif stype in ("Product","Service","LocalBusiness","ProfessionalService","HomeAndConstructionBusiness"):
            for field in ("name","description","serviceType","category","areaServed"):
                val = obj.get(field,"")
                text = get_text(val) if isinstance(val,dict) else clean_text(val) if isinstance(val,str) else ""
                if isinstance(val, list):
                    text = " | ".join(clean_text(v) if isinstance(v,str) else get_text(v) for v in val)
                if text and text not in seen_q and len(text)>3:
                    seen_q.add(text)
                    pairs.append({"question":text,"answer":"","schema_type":stype,"source_field":field})

        elif stype == "BreadcrumbList":
            items = obj.get("itemListElement",[])
            if isinstance(items, dict): items = [items]
            for item in (items if isinstance(items, list) else []):
                name = clean_text(item.get("name",""))
                if name and name not in seen_q and len(name.split())>=2:
                    seen_q.add(name)
                    pairs.append({"question":name,"answer":"","schema_type":"Breadcrumb","source_field":"item.name"})

        for key in ("@graph","hasPart","mainEntity","mainEntityOfPage","about","mentions"):
            child = obj.get(key)
            if child: walk(child, stype)

    for data in parsed_blocks:
        walk(data)
    return pairs

def fetch_page_content(url, timeout=12):
    try:
        h = BASE_HEADERS.copy()
        h["Accept-Encoding"] = "gzip, deflate"
        r = requests.get(url, headers=h, timeout=timeout, allow_redirects=True)
        if r.status_code != 200:
            return None
        ct = r.headers.get("content-type","")
        if "html" not in ct and "text" not in ct:
            return None
        html = r.text[:300_000]

        # Fix M: use the final redirected URL as canonical URL
        final_url = r.url if r.url else url
        domain = extract_domain(final_url)

        # Fix Q: parse JSON-LD once, share with both extractors
        parsed_blocks = _parse_jsonld_blocks(html)
        schema_texts  = extract_jsonld_texts(parsed_blocks)
        schema_paa    = extract_schema_paa(parsed_blocks)

        html = re.sub(r'<(script|style|nav|footer|header|aside|noscript)[^>]*>.*?</\1>','',html,flags=re.DOTALL|re.IGNORECASE)
        html = re.sub(r'<!--.*?-->','',html,flags=re.DOTALL)

        def clean(s):
            s = re.sub(r'<[^>]+>','',s)
            s = re.sub(r'&[a-z]+;',' ',s)
            s = re.sub(r'\s+',' ',s).strip()
            return s

        title_raw = re.search(r'<title[^>]*>(.*?)</title>',html,re.IGNORECASE|re.DOTALL)
        meta_desc = (re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']{10,400})["\']',html,re.IGNORECASE) or
                     re.search(r'<meta[^>]+content=["\']([^"\']{10,400})["\'][^>]+name=["\']description["\']',html,re.IGNORECASE))
        meta_kw   = (re.search(r'<meta[^>]+name=["\']keywords["\'][^>]+content=["\']([^"\']{5,400})["\']',html,re.IGNORECASE) or
                     re.search(r'<meta[^>]+content=["\']([^"\']{5,400})["\'][^>]+name=["\']keywords["\']',html,re.IGNORECASE))
        h1s = [clean(m) for m in re.findall(r'<h1[^>]*>(.*?)</h1>',html,re.IGNORECASE|re.DOTALL)]
        h2s = [clean(m) for m in re.findall(r'<h2[^>]*>(.*?)</h2>',html,re.IGNORECASE|re.DOTALL)]
        h3s = [clean(m) for m in re.findall(r'<h3[^>]*>(.*?)</h3>',html,re.IGNORECASE|re.DOTALL)]
        alts = re.findall(r'alt=["\']([^"\']{5,120})["\']',html,re.IGNORECASE)
        body_chunks = [clean(m) for m in re.findall(r'<p[^>]*>(.*?)</p>',html,re.IGNORECASE|re.DOTALL)]
        body_text = ' '.join(body_chunks)[:50_000]

        return {
            "domain":domain,"url":final_url,
            "title":  clean(title_raw.group(1)) if title_raw else "",
            "meta":   clean(meta_desc.group(1)) if meta_desc else "",
            "meta_kw":clean(meta_kw.group(1))   if meta_kw  else "",
            "h1":h1s[:5],"h2":h2s[:20],"h3":h3s[:30],
            "alts":alts[:20],"body":body_text,
            "schema":schema_texts,"schema_paa":schema_paa,
        }
    except Exception as e:
        print(f"Page fetch error {url}: {e}")
        return None

def _zone_label(weight):
    if weight >= 8: return "title"
    if weight >= 7: return "h1"
    if weight >= 6: return "meta"
    if weight >= 5: return "h2"
    if weight >= 3: return "h3"
    return "body"

def extract_page_keywords(page, seed_words, min_len=3, max_len=7):
    seed_words_set = set(w.lower() for w in seed_words)
    zones = []
    if page["title"]:    zones.append((page["title"], 8))
    if page["meta"]:     zones.append((page["meta"], 6))
    if page["meta_kw"]:  zones.append((page["meta_kw"], 7))
    for h in page["h1"]: zones.append((h, 7))
    for h in page["h2"]: zones.append((h, 5))
    for h in page["h3"]: zones.append((h, 3))
    for a in page["alts"]: zones.append((a, 3))
    zones.append((page["body"], 1))
    schema_entries = page.get("schema", [])

    ngram_scores = Counter()
    # Fix H: track highest-weight zone per phrase, not first-seen
    ngram_zone_weight = {}
    ngram_zone        = {}

    for text, weight in zones:
        if not text: continue
        tokens = re.findall(r"[a-z][a-z'\-]{1,30}", text.lower())
        tokens = [t.strip("'-") for t in tokens]
        n = len(tokens)
        # Fix E: removed dead word_count variable; gram_len already constrains range
        for gram_len in range(2, 6):
            for i in range(n - gram_len + 1):
                gram = tokens[i:i+gram_len]
                if gram[0] in STOP_WORDS or gram[-1] in STOP_WORDS: continue
                if all(w in STOP_WORDS for w in gram): continue
                phrase = ' '.join(gram)
                if len(phrase) < min_len or len(phrase) > 80: continue
                seed_bonus = 2 if any(sw in gram for sw in seed_words_set) else 1
                ngram_scores[phrase] += weight * seed_bonus
                if weight > ngram_zone_weight.get(phrase, -1):
                    ngram_zone_weight[phrase] = weight
                    ngram_zone[phrase] = _zone_label(weight)

    for text, weight, zone_label in schema_entries:
        tokens = re.findall(r"[a-z][a-z'\-]{1,30}", text.lower())
        tokens = [t.strip("'-") for t in tokens]
        n = len(tokens)
        if 1 <= n <= max_len:
            phrase = ' '.join(t for t in tokens if t)
            phrase = re.sub(r'\s+',' ', phrase).strip()
            words = phrase.split()
            if len(words) >= 2 and words[0] not in STOP_WORDS and words[-1] not in STOP_WORDS:
                seed_bonus = 2 if any(sw in tokens for sw in seed_words_set) else 1
                ngram_scores[phrase] += weight * seed_bonus * 1.5
                if weight > ngram_zone_weight.get(phrase, -1):
                    ngram_zone_weight[phrase] = weight
                    ngram_zone[phrase] = zone_label
        for gram_len in range(2, 6):
            for i in range(n - gram_len + 1):
                gram = tokens[i:i+gram_len]
                if gram[0] in STOP_WORDS or gram[-1] in STOP_WORDS: continue
                if all(w in STOP_WORDS for w in gram): continue
                phrase = ' '.join(gram)
                if len(phrase) < min_len or len(phrase) > 80: continue
                seed_bonus = 2 if any(sw in gram for sw in seed_words_set) else 1
                ngram_scores[phrase] += weight * seed_bonus * 1.5
                if weight > ngram_zone_weight.get(phrase, -1):
                    ngram_zone_weight[phrase] = weight
                    ngram_zone[phrase] = zone_label

    for h in (page["h1"] + page["h2"] + page["h3"]):
        phrase = re.sub(r'[^a-z0-9 \-]','',h.lower()).strip()
        phrase = re.sub(r'\s+',' ',phrase)
        words  = phrase.split()
        if 2 <= len(words) <= max_len and words[0] not in STOP_WORDS and words[-1] not in STOP_WORDS:
            ngram_scores[phrase] += 10
            if 10 > ngram_zone_weight.get(phrase, -1):
                ngram_zone_weight[phrase] = 10
                ngram_zone[phrase] = "heading"

    results = []
    for phrase, score in ngram_scores.most_common(300):
        if score < 2: break
        if len(phrase.split()) < 2: continue
        results.append({"keyword":phrase,"score":round(score,1),
                        "zone":ngram_zone.get(phrase,"body"),"source":page["domain"]})
    return results[:150]

def get_reddit_results(query):
    results=[]; seen=set()
    try:
        r=requests.get("https://www.reddit.com/search.json",
            params={"q":query,"limit":25,"sort":"relevance","type":"link"},
            headers=REDDIT_HEADERS, timeout=10)
        for post in r.json().get("data",{}).get("children",[]):
            d=post.get("data",{})
            title=d.get("title","").strip(); sub=d.get("subreddit","").strip()
            if title and title not in seen:
                seen.add(title); results.append({"text":title,"source":f"r/{sub}","type":"post"})
        time.sleep(0.5)
        r2=requests.get("https://www.reddit.com/search.json",
            params={"q":query,"limit":10,"sort":"relevance","type":"sr"},
            headers=REDDIT_HEADERS, timeout=10)
        for sub in r2.json().get("data",{}).get("children",[]):
            d=sub.get("data",{})
            title=d.get("title","").strip(); name=d.get("display_name_prefixed","")
            if title and title not in seen:
                seen.add(title); results.append({"text":title,"source":name,"type":"subreddit"})
    except Exception as e: print(f"Reddit error: {e}")
    return results[:30]

def get_quora_results(query):
    results=[]; seen=set()
    try:
        url=f"https://www.quora.com/search?q={requests.utils.quote(query)}&type=question"
        r=requests.get(url, headers={**BASE_HEADERS,"Referer":"https://www.quora.com/"}, timeout=12)
        html=r.text
        for _,title in re.findall(r'"question":\{"url":"([^"]+)","title":"([^"]+)"', html):
            title=title.strip()
            if title and title not in seen and len(title)>10:
                seen.add(title); results.append({"text":title,"source":"quora.com","type":"question"})
        if not results:
            for a in re.findall(r'<a[^>]+class="[^"]*qu-[^"]*"[^>]*>([^<]{15,150})</a>', html):
                a=a.strip()
                if a and '?' in a and a not in seen:
                    seen.add(a); results.append({"text":a,"source":"quora.com","type":"question"})
    except Exception as e: print(f"Quora error: {e}")
    return results[:25]

def build_queries(keyword, mode, chars, modifiers):
    if mode=="modifiers":
        return [f"{m} * {keyword}" for m in modifiers if m.strip()]
    elif mode=="modifiers_alpha":
        return [f"{m} * {keyword}" for m in modifiers if m.strip()] + [f"{keyword} {c}" for c in chars]
    else:
        return [f"{keyword} {c}" for c in chars]

def make_ac_route(fetch_fn):
    def handler():
        seed      = request.args.get("seed","").strip()
        depth     = int(request.args.get("depth",1))
        mode      = request.args.get("mode","alpha")
        modifiers = request.args.get("modifiers","").split("|")
        location  = request.args.get("location","").strip()
        gl        = request.args.get("gl","us")
        hl        = GL_MAP.get(gl,"en")
        min_words = int(request.args.get("min_words",0))
        max_words = int(request.args.get("max_words",0))
        exclude   = [x.strip().lower() for x in request.args.get("exclude","").split(",") if x.strip()]
        if not seed: return jsonify({"error":"No seed"}), 400
        full_seed = f"{seed} {location}".strip() if location else seed
        chars = NUMBERS if mode=="numeric" else (ALPHANUM if mode=="both" else ALPHABET)
        def generate():
            seen=set(); queue=[(full_seed,0)]; total=0
            while queue:
                keyword,current_depth=queue.pop(0)
                if current_depth>=depth: continue
                queries=build_queries(keyword,mode,chars,modifiers)
                for i,query in enumerate(queries):
                    if fetch_fn==get_autocomplete:
                        results=fetch_fn(query,hl,gl)
                    else:
                        results=fetch_fn(query)
                    for s in results:
                        if not isinstance(s, str): continue
                        s=s.strip().lower()
                        if not s or s in seen: continue
                        wc=len(s.split())
                        if min_words and wc<min_words: continue
                        if max_words and wc>max_words: continue
                        if any(ex in s for ex in exclude): continue
                        seen.add(s); total+=1
                        yield f"data: {json.dumps({'keyword':s,'parent':keyword,'depth':current_depth+1,'total':total})}\n\n"
                        if current_depth+1<depth: queue.append((s,current_depth+1))
                    yield f"data: {json.dumps({'progress':True,'query':query,'index':i+1,'of':len(queries),'total':total})}\n\n"
                    time.sleep(0.3)
            yield "data: {\"done\": true}\n\n"
        return Response(generate(), mimetype="text/event-stream",
                        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})
    return handler

research_autocomplete = make_ac_route(get_autocomplete)
research_bing         = make_ac_route(get_bing_autocomplete)
research_youtube      = make_ac_route(get_youtube_autocomplete)
app.add_url_rule("/research/autocomplete", "research_autocomplete", research_autocomplete)
app.add_url_rule("/research/bing",         "research_bing",         research_bing)
app.add_url_rule("/research/youtube",      "research_youtube",      research_youtube)

# Fix L: None default instead of mutable [] default
def stream_scrape(seed, fetch_fn, limit=15, location="", gl="us", min_words=0, max_words=0, exclude=None):
    if exclude is None: exclude = []
    hl = GL_MAP.get(gl,"en")
    full_seed = f"{seed} {location}".strip() if location else seed
    clear_google_cache()
    def generate():
        seen=set(); queue=[full_seed]; total=0; checked=0
        while queue and checked<limit:
            keyword=queue.pop(0); checked+=1
            results=fetch_fn(keyword,hl,gl)
            yield f"data: {json.dumps({'progress':True,'query':keyword,'index':checked,'of':limit,'total':total})}\n\n"
            for r in results:
                r=r.strip()
                if not r or r in seen: continue
                wc=len(r.split())
                if min_words and wc<min_words: continue
                if max_words and wc>max_words: continue
                if any(ex in r.lower() for ex in exclude): continue
                seen.add(r); total+=1
                yield f"data: {json.dumps({'keyword':r,'parent':keyword,'depth':1,'total':total})}\n\n"
                queue.append(r)
            time.sleep(1)
        yield "data: {\"done\": true}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/")
def index(): return send_from_directory("static","index.html")

@app.route("/research/paa")
def research_paa():
    seed=request.args.get("seed","").strip()
    if not seed: return jsonify({"error":"No seed"}),400
    return stream_scrape(seed,get_people_also_ask,20,
        request.args.get("location",""),request.args.get("gl","us"),
        int(request.args.get("min_words",0)),int(request.args.get("max_words",0)),
        [x.strip().lower() for x in request.args.get("exclude","").split(",") if x.strip()])

@app.route("/research/related")
def research_related():
    seed=request.args.get("seed","").strip()
    if not seed: return jsonify({"error":"No seed"}),400
    return stream_scrape(seed,get_related_searches,15,
        request.args.get("location",""),request.args.get("gl","us"),
        int(request.args.get("min_words",0)),int(request.args.get("max_words",0)),
        [x.strip().lower() for x in request.args.get("exclude","").split(",") if x.strip()])

@app.route("/research/pasf")
def research_pasf():
    seed=request.args.get("seed","").strip()
    if not seed: return jsonify({"error":"No seed"}),400
    return stream_scrape(seed,get_people_also_search,15,
        request.args.get("location",""),request.args.get("gl","us"),
        int(request.args.get("min_words",0)),int(request.args.get("max_words",0)),
        [x.strip().lower() for x in request.args.get("exclude","").split(",") if x.strip()])

@app.route("/research/local")
def research_local():
    seed=request.args.get("seed","").strip()
    location=request.args.get("location","").strip()
    gl=request.args.get("gl","us"); hl=GL_MAP.get(gl,"en")
    if not seed: return jsonify({"error":"No seed"}),400
    full_seed=f"{seed} {location}".strip() if location else seed
    def generate():
        yield f"data: {json.dumps({'progress':True,'query':full_seed,'index':1,'of':1,'total':0})}\n\n"
        for i,biz in enumerate(get_local_pack(full_seed,hl,gl)):
            yield f"data: {json.dumps({'business':biz,'index':i,'total':3})}\n\n"
            time.sleep(0.2)
        yield "data: {\"done\": true}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/research/reddit")
def research_reddit():
    seed=request.args.get("seed","").strip()
    location=request.args.get("location","").strip()
    if not seed: return jsonify({"error":"No seed"}),400
    full_seed=f"{seed} {location}".strip() if location else seed
    min_words=int(request.args.get("min_words",0)); max_words=int(request.args.get("max_words",0))
    exclude=[x.strip().lower() for x in request.args.get("exclude","").split(",") if x.strip()]
    def generate():
        yield f"data: {json.dumps({'progress':True,'query':full_seed,'index':1,'of':1,'total':0})}\n\n"
        results=get_reddit_results(full_seed); total=0
        for r in results:
            wc=len(r['text'].split())
            if min_words and wc<min_words: continue
            if max_words and wc>max_words: continue
            if any(ex in r['text'].lower() for ex in exclude): continue
            total+=1
            yield f"data: {json.dumps({'keyword':r['text'],'source':r['source'],'type':r['type'],'depth':1,'total':total})}\n\n"
        yield "data: {\"done\": true}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/research/quora")
def research_quora():
    seed=request.args.get("seed","").strip()
    location=request.args.get("location","").strip()
    if not seed: return jsonify({"error":"No seed"}),400
    full_seed=f"{seed} {location}".strip() if location else seed
    min_words=int(request.args.get("min_words",0)); max_words=int(request.args.get("max_words",0))
    exclude=[x.strip().lower() for x in request.args.get("exclude","").split(",") if x.strip()]
    def generate():
        yield f"data: {json.dumps({'progress':True,'query':full_seed,'index':1,'of':1,'total':0})}\n\n"
        results=get_quora_results(full_seed); total=0
        for r in results:
            wc=len(r['text'].split())
            if min_words and wc<min_words: continue
            if max_words and wc>max_words: continue
            if any(ex in r['text'].lower() for ex in exclude): continue
            total+=1
            yield f"data: {json.dumps({'keyword':r['text'],'source':r['source'],'type':r['type'],'depth':1,'total':total})}\n\n"
        yield "data: {\"done\": true}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/research/serp", methods=["POST"])
def research_serp():
    data = request.json
    keywords = data.get("keywords",[])[:200]
    engine   = data.get("engine","both")
    gl       = data.get("gl","us")
    hl       = GL_MAP.get(gl,"en")
    if not keywords: return jsonify({"error":"No keywords"}),400
    pairs=[]
    for kw in keywords:
        if engine in ("google","both"): pairs.append((kw,"google"))
        if engine in ("bing","both"):   pairs.append((kw,"bing"))
    def generate():
        total=len(pairs)
        for i,(kw,eng) in enumerate(pairs):
            urls = get_google_serp_urls(kw,hl,gl) if eng=="google" else get_bing_serp_urls(kw)
            enriched=[{"rank":r+1,"url":u,"domain":extract_domain(u)} for r,u in enumerate(urls)]
            yield f"data: {json.dumps({'keyword':kw,'engine':eng,'urls':enriched,'index':i+1,'total':total})}\n\n"
            time.sleep(0.8)
        yield "data: {\"done\": true}\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/research/page_keywords", methods=["POST"])
def research_page_keywords():
    data     = request.json
    urls     = data.get("urls",[])[:50]
    seed     = data.get("seed","")
    existing = set(k.lower().strip() for k in data.get("existing_keywords",[]))
    seed_words = [w for w in re.findall(r'[a-z]+', seed.lower()) if w not in STOP_WORDS and len(w)>2]
    if not urls: return jsonify({"error":"No URLs provided"}),400

    def generate():
        total=len(urls)
        all_seen=set(); crawled_urls=set()
        for i,url in enumerate(urls):
            norm=re.sub(r'^(https?://[^/]+)', lambda m: m.group(1).lower(), url).rstrip('/')
            if norm in crawled_urls:
                yield f"data: {json.dumps({'progress':True,'url':url,'index':i+1,'total':total,'skipped':True})}\n\n"
                continue
            crawled_urls.add(norm)
            yield f"data: {json.dumps({'progress':True,'url':url,'index':i+1,'total':total})}\n\n"
            page=fetch_page_content(url)
            if page is None:
                yield f"data: {json.dumps({'error':True,'url':url,'message':'Could not fetch page','index':i+1,'total':total})}\n\n"
                time.sleep(0.3); continue
            raw_kws=extract_page_keywords(page,seed_words)
            fresh=[]
            for entry in raw_kws:
                kw=entry["keyword"].lower().strip()
                if kw in all_seen: continue
                all_seen.add(kw); entry["is_new"]=kw not in existing; fresh.append(entry)
            raw_paa=page.get("schema_paa",[])
            fresh_paa=[]
            for entry in raw_paa:
                q=entry["question"].lower().strip()
                if q in all_seen: continue
                all_seen.add(q); entry["is_new"]=q not in existing; fresh_paa.append(entry)
            # Fix M: use page['url'] which is the final redirected URL
            yield f"data: {json.dumps({'url':page['url'],'domain':page['domain'],'title':page['title'],'keywords':fresh,'schema_paa':fresh_paa,'index':i+1,'total':total})}\n\n"
            time.sleep(1.0)
        yield "data: {\"done\": true}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/analyze/wordfreq", methods=["POST"])
def word_freq():
    data=request.json; all_keywords=[]
    for kws in data.get("modules",{}).values():
        all_keywords.extend(kws)
    words=[]
    for kw in all_keywords:
        for w in kw.lower().split():
            w=re.sub(r'[^a-z0-9]','',w)
            if w and len(w)>2 and w not in STOP_WORDS:
                words.append(w)
    return jsonify({"freq":Counter(words).most_common(50)})

@app.route("/export/xlsx", methods=["POST"])
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error":"openpyxl not installed. Run: pip install openpyxl"}),500

    data=request.json; seed=data.get("seed","keywords")
    keyword_sheets=[
        ("Google Autocomplete",    data.get("ac",[]),     "00B4CC"),
        ("Bing Autocomplete",      data.get("bing",[]),   "008272"),
        ("YouTube Autocomplete",   data.get("youtube",[]),"FF0000"),
        ("People Also Ask",        data.get("paa",[]),    "F59E0B"),
        ("Related Searches",       data.get("rel",[]),    "06B6D4"),
        ("People Also Search For", data.get("pasf",[]),   "EC4899"),
        ("Reddit",                 data.get("reddit",[]), "FF4500"),
        ("Quora",                  data.get("quora",[]),  "B92B27"),
    ]
    wb=Workbook(); wb.remove(wb.active)

    all_kws=[]; seen_all=set()
    for _,kws,_ in keyword_sheets:
        for kw in kws:
            kw=kw.strip().lower()
            if kw and kw not in seen_all: seen_all.add(kw); all_kws.append(kw)

    if all_kws:
        ws=wb.create_sheet(title="All Keywords")
        ws.column_dimensions["A"].width=55
        h=ws.cell(row=1,column=1,value="Keyword")
        h.font=Font(bold=True,color="FFFFFF",size=11); h.fill=PatternFill("solid",fgColor="374151"); h.alignment=Alignment(horizontal="left")
        for i,kw in enumerate(sorted(all_kws),start=2):
            cell=ws.cell(row=i,column=1,value=kw); cell.alignment=Alignment(horizontal="left")
            if i%2==0: cell.fill=PatternFill("solid",fgColor="F8F9FA")
        ws.cell(row=len(all_kws)+3,column=1,value=f"Total unique: {len(all_kws)}").font=Font(italic=True,color="888888")

    word_data=data.get("wordfreq",[])
    if word_data:
        ws=wb.create_sheet(title="Word Frequency")
        ws.column_dimensions["A"].width=25; ws.column_dimensions["B"].width=15
        for col,hdr in enumerate(["Word","Count"],start=1):
            cell=ws.cell(row=1,column=col,value=hdr)
            cell.font=Font(bold=True,color="FFFFFF",size=11); cell.fill=PatternFill("solid",fgColor="374151"); cell.alignment=Alignment(horizontal="left")
        for i,(word,count) in enumerate(word_data,start=2):
            ws.cell(row=i,column=1,value=word).alignment=Alignment(horizontal="left")
            ws.cell(row=i,column=2,value=count)
            if i%2==0:
                ws.cell(row=i,column=1).fill=PatternFill("solid",fgColor="F8F9FA")
                ws.cell(row=i,column=2).fill=PatternFill("solid",fgColor="F8F9FA")

    for sheet_name,keywords,color in keyword_sheets:
        if not keywords: continue
        ws=wb.create_sheet(title=sheet_name); ws.column_dimensions["A"].width=55
        h=ws.cell(row=1,column=1,value="Keyword")
        h.font=Font(bold=True,color="FFFFFF",size=11); h.fill=PatternFill("solid",fgColor=color); h.alignment=Alignment(horizontal="left")
        seen=set(); unique=[]
        for kw in keywords:
            kw=kw.strip().lower()
            if kw and kw not in seen: seen.add(kw); unique.append(kw)
        for i,kw in enumerate(unique,start=2):
            cell=ws.cell(row=i,column=1,value=kw); cell.alignment=Alignment(horizontal="left")
            if i%2==0: cell.fill=PatternFill("solid",fgColor="F8F9FA")
        ws.cell(row=len(unique)+3,column=1,value=f"Total: {len(unique)} keywords").font=Font(italic=True,color="888888")

    local=data.get("local",[])
    if local:
        ws=wb.create_sheet(title="Local Pack")
        headers=["Business Name","Rating","Reviews","Phone","Address","Category","Source Query"]
        widths=[35,10,10,18,40,25,30]
        for col,(hdr,w) in enumerate(zip(headers,widths),start=1):
            ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
            cell=ws.cell(row=1,column=col,value=hdr)
            cell.font=Font(bold=True,color="FFFFFF",size=11); cell.fill=PatternFill("solid",fgColor="1A73E8"); cell.alignment=Alignment(horizontal="left")
        for i,biz in enumerate(local,start=2):
            for col,val in enumerate([biz.get("name",""),biz.get("rating",""),biz.get("reviews",""),biz.get("phone",""),biz.get("address",""),biz.get("category",""),biz.get("query","")],start=1):
                cell=ws.cell(row=i,column=col,value=val); cell.alignment=Alignment(horizontal="left")
                if i%2==0: cell.fill=PatternFill("solid",fgColor="EBF5FB")

    for serp_key,serp_title,serp_color in [("serp_google","Google SERP URLs","00B4CC"),("serp_bing","Bing SERP URLs","008272")]:
        serp_data=data.get(serp_key,[])
        if not serp_data: continue
        ws=wb.create_sheet(title=serp_title)
        headers=["Keyword","Rank","URL","Domain"]; widths=[45,6,70,35]
        for col,(hdr,w) in enumerate(zip(headers,widths),start=1):
            ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
            cell=ws.cell(row=1,column=col,value=hdr)
            cell.font=Font(bold=True,color="FFFFFF",size=11); cell.fill=PatternFill("solid",fgColor=serp_color); cell.alignment=Alignment(horizontal="left")
        row=2
        for entry in serp_data:
            kw=entry.get("keyword","")
            for u in entry.get("urls",[]):
                for col,val in enumerate([kw,u.get("rank",""),u.get("url",""),u.get("domain","")],start=1):
                    cell=ws.cell(row=row,column=col,value=val); cell.alignment=Alignment(horizontal="left")
                    if row%2==0: cell.fill=PatternFill("solid",fgColor="EBF5FB" if serp_key=="serp_bing" else "E6F7FA")
                    if col==3:
                        try: cell.hyperlink=val; cell.font=Font(color="0563C1",underline="single")
                        except: pass
                row+=1
        ws.cell(row=row+1,column=1,value=f"Total: {row-2} URL entries").font=Font(italic=True,color="888888")

    page_kw_data=data.get("page_keywords",[])
    if page_kw_data:
        ws=wb.create_sheet(title="Page Keywords")
        headers=["Keyword","Score","Zone","Source URL","Domain","New?"]; widths=[50,8,10,70,30,8]
        for col,(hdr,w) in enumerate(zip(headers,widths),start=1):
            ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
            cell=ws.cell(row=1,column=col,value=hdr)
            cell.font=Font(bold=True,color="FFFFFF",size=11); cell.fill=PatternFill("solid",fgColor="166534"); cell.alignment=Alignment(horizontal="left")
        row=2
        for entry in page_kw_data:
            is_new=entry.get("is_new",False)
            for col,val in enumerate([entry.get("keyword",""),entry.get("score",""),entry.get("zone",""),entry.get("url",""),entry.get("source",""),"✓ NEW" if is_new else ""],start=1):
                cell=ws.cell(row=row,column=col,value=val); cell.alignment=Alignment(horizontal="left")
                if is_new: cell.fill=PatternFill("solid",fgColor="D1FAE5")
                elif row%2==0: cell.fill=PatternFill("solid",fgColor="F8F9FA")
            row+=1
        ws.cell(row=row+1,column=1,value=f"Total: {row-2} page keyword entries").font=Font(italic=True,color="888888")

    schema_paa_data=data.get("schema_paa",[])
    if schema_paa_data:
        ws=wb.create_sheet(title="Schema PAA")
        headers=["Question","Answer","Schema Type","Source Field","URL","Domain","New?"]; widths=[60,80,20,18,60,30,8]
        for col,(hdr,w) in enumerate(zip(headers,widths),start=1):
            ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
            cell=ws.cell(row=1,column=col,value=hdr)
            cell.font=Font(bold=True,color="FFFFFF",size=11); cell.fill=PatternFill("solid",fgColor="D97706"); cell.alignment=Alignment(horizontal="left",wrap_text=True)
        row=2
        for entry in schema_paa_data:
            is_new=entry.get("is_new",False)
            vals=[entry.get("question",""),entry.get("answer",""),entry.get("schema_type",""),entry.get("source_field",""),entry.get("url",""),entry.get("domain",""),"✓ NEW" if is_new else ""]
            for col,val in enumerate(vals,start=1):
                cell=ws.cell(row=row,column=col,value=val); cell.alignment=Alignment(horizontal="left",wrap_text=True)
                if is_new: cell.fill=PatternFill("solid",fgColor="FEF3C7")
                elif row%2==0: cell.fill=PatternFill("solid",fgColor="F8F9FA")
            row+=1

    output=io.BytesIO(); wb.save(output); output.seek(0)
    slug=seed.replace(" ","-").lower()
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"{slug}-keyword-research.xlsx")

if __name__=="__main__":
    os.makedirs("static",exist_ok=True)
    import os
app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)),debug=False)
